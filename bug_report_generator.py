import logging
import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# ─── Configuration du logger ───────────────────────────────────────────────────

os.makedirs("reports", exist_ok=True)
os.makedirs("logs", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("logs/bug_report.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ─── Données simulées de tests ─────────────────────────────────────────────────

SIMULATED_TEST_RESULTS = [
    {"id": "TC-001", "module": "Login",        "description": "Valid user login",               "status": "PASS", "duration": 1.2},
    {"id": "TC-002", "module": "Login",        "description": "Empty password field",           "status": "FAIL", "duration": 0.8},
    {"id": "TC-003", "module": "Login",        "description": "Invalid email format",           "status": "FAIL", "duration": 0.5},
    {"id": "TC-004", "module": "Dashboard",    "description": "Dashboard loads correctly",      "status": "PASS", "duration": 2.1},
    {"id": "TC-005", "module": "Dashboard",    "description": "Statistics chart renders",       "status": "PASS", "duration": 1.7},
    {"id": "TC-006", "module": "Search",       "description": "Search returns results",         "status": "PASS", "duration": 0.9},
    {"id": "TC-007", "module": "Search",       "description": "Search with special characters", "status": "FAIL", "duration": 1.1},
    {"id": "TC-008", "module": "Search",       "description": "Empty search field",             "status": "PASS", "duration": 0.4},
    {"id": "TC-009", "module": "Profile",      "description": "Update user profile",            "status": "FAIL", "duration": 2.3},
    {"id": "TC-010", "module": "Profile",      "description": "Upload profile picture",         "status": "PASS", "duration": 3.1},
    {"id": "TC-011", "module": "Profile",      "description": "Change password",                "status": "PASS", "duration": 1.5},
    {"id": "TC-012", "module": "Checkout",     "description": "Add item to cart",               "status": "PASS", "duration": 0.7},
    {"id": "TC-013", "module": "Checkout",     "description": "Apply discount code",            "status": "FAIL", "duration": 1.0},
    {"id": "TC-014", "module": "Checkout",     "description": "Payment with credit card",       "status": "PASS", "duration": 4.2},
    {"id": "TC-015", "module": "Checkout",     "description": "Order confirmation email",       "status": "PASS", "duration": 2.8},
    {"id": "TC-016", "module": "API",          "description": "GET /users returns 200",         "status": "PASS", "duration": 0.3},
    {"id": "TC-017", "module": "API",          "description": "POST /users returns 201",        "status": "PASS", "duration": 0.4},
    {"id": "TC-018", "module": "API",          "description": "DELETE /users/999 returns 404",  "status": "FAIL", "duration": 0.6},
    {"id": "TC-019", "module": "Performance",  "description": "Homepage load under 3s",         "status": "PASS", "duration": 2.4},
    {"id": "TC-020", "module": "Performance",  "description": "Search response under 1s",       "status": "FAIL", "duration": 1.8},
]

SEVERITIES = {
    "Login":       "Critical",
    "Checkout":    "High",
    "API":         "High",
    "Search":      "Medium",
    "Profile":     "Medium",
    "Dashboard":   "Low",
    "Performance": "Low",
}

BUG_DETAILS = {
    "TC-002": ("Login button unresponsive when password is empty",      "Add validation message for empty password field"),
    "TC-003": ("No error shown for invalid email format",               "Implement regex validation on email input"),
    "TC-007": ("App crashes with special chars in search bar",          "Sanitize input to handle special characters"),
    "TC-009": ("Profile update fails silently — no confirmation shown", "Fix API call and add success/error toast"),
    "TC-013": ("Discount code field throws 500 error",                  "Fix backend discount validation endpoint"),
    "TC-018": ("DELETE on unknown ID returns 200 instead of 404",       "Fix HTTP status code in delete endpoint"),
    "TC-020": ("Search response exceeds 1s threshold (1.8s measured)",  "Optimize search query and add caching"),
}


# ─── Générateur de rapport ─────────────────────────────────────────────────────

def generate_bug_report():
    logger.info("=" * 60)
    logger.info("BUG REPORT GENERATOR — Started")
    logger.info("=" * 60)

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Résumé ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "📊 Summary"

    total   = len(SIMULATED_TEST_RESULTS)
    passed  = sum(1 for t in SIMULATED_TEST_RESULTS if t["status"] == "PASS")
    failed  = sum(1 for t in SIMULATED_TEST_RESULTS if t["status"] == "FAIL")
    pass_rate = round((passed / total) * 100, 1)

    logger.info(f"Total tests : {total} | Passed : {passed} | Failed : {failed} | Pass rate : {pass_rate}%")

    # Couleurs
    green  = PatternFill("solid", fgColor="2ECC71")
    red    = PatternFill("solid", fgColor="E74C3C")
    blue   = PatternFill("solid", fgColor="2980B9")
    orange = PatternFill("solid", fgColor="E67E22")
    grey   = PatternFill("solid", fgColor="BDC3C7")
    white_bg = PatternFill("solid", fgColor="FFFFFF")
    light_grey = PatternFill("solid", fgColor="F2F3F4")

    bold_white = Font(bold=True, color="FFFFFF", size=11)
    bold_dark  = Font(bold=True, color="2C3E50", size=11)
    center     = Alignment(horizontal="center", vertical="center")

    def border():
        side = Side(style="thin", color="BDC3C7")
        return Border(left=side, right=side, top=side, bottom=side)

    # Titre
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = "🐛 AUTOMATED BUG REPORT"
    ws_summary["A1"].font = Font(bold=True, size=16, color="2C3E50")
    ws_summary["A1"].alignment = center
    ws_summary["A1"].fill = light_grey

    ws_summary["A2"] = f"Generated on : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = Font(italic=True, color="7F8C8D")
    ws_summary.merge_cells("A2:D2")
    ws_summary["A2"].alignment = center

    ws_summary.append([])

    # KPIs
    kpis = [
        ("Total Tests",  total,     blue),
        ("✅ Passed",    passed,    green),
        ("❌ Failed",    failed,    red),
        (f"Pass Rate",   f"{pass_rate}%", orange),
    ]

    headers = [k[0] for k in kpis]
    values  = [k[1] for k in kpis]
    fills   = [k[2] for k in kpis]

    ws_summary.append(headers)
    ws_summary.append(values)

    header_row = 4
    value_row  = 5

    for col_idx, (fill, val) in enumerate(zip(fills, values), start=1):
        h_cell = ws_summary.cell(row=header_row, column=col_idx)
        v_cell = ws_summary.cell(row=value_row,  column=col_idx)
        h_cell.fill = fill
        h_cell.font = bold_white
        h_cell.alignment = center
        h_cell.border = border()
        v_cell.font = Font(bold=True, size=14, color="2C3E50")
        v_cell.alignment = center
        v_cell.border = border()

    ws_summary.append([])

    # Résultats par module
    ws_summary.append(["Module", "Total", "Passed", "Failed"])
    mod_header_row = ws_summary.max_row
    for col in range(1, 5):
        cell = ws_summary.cell(row=mod_header_row, column=col)
        cell.fill = blue
        cell.font = bold_white
        cell.alignment = center
        cell.border = border()

    modules = {}
    for t in SIMULATED_TEST_RESULTS:
        m = t["module"]
        modules.setdefault(m, {"total": 0, "pass": 0, "fail": 0})
        modules[m]["total"] += 1
        if t["status"] == "PASS":
            modules[m]["pass"] += 1
        else:
            modules[m]["fail"] += 1

    for mod, stats in modules.items():
        ws_summary.append([mod, stats["total"], stats["pass"], stats["fail"]])
        row = ws_summary.max_row
        ws_summary.cell(row=row, column=1).font = bold_dark
        for col in range(1, 5):
            cell = ws_summary.cell(row=row, column=col)
            cell.alignment = center
            cell.border = border()
            if col == 4 and stats["fail"] > 0:
                cell.fill = PatternFill("solid", fgColor="FADBD8")
                cell.font = Font(bold=True, color="C0392B")

    for col in range(1, 5):
        ws_summary.column_dimensions[get_column_letter(col)].width = 20

    ws_summary.row_dimensions[1].height = 35
    ws_summary.row_dimensions[5].height = 30

    # ── Feuille 2 : Tous les tests ──────────────────────────────────────────────
    ws_tests = wb.create_sheet("🧪 All Tests")
    headers_tests = ["Test ID", "Module", "Description", "Status", "Duration (s)"]
    ws_tests.append(headers_tests)

    for col in range(1, 6):
        cell = ws_tests.cell(row=1, column=col)
        cell.fill = blue
        cell.font = bold_white
        cell.alignment = center
        cell.border = border()

    for t in SIMULATED_TEST_RESULTS:
        ws_tests.append([t["id"], t["module"], t["description"], t["status"], t["duration"]])
        row = ws_tests.max_row
        status_cell = ws_tests.cell(row=row, column=4)
        if t["status"] == "PASS":
            status_cell.fill = PatternFill("solid", fgColor="D5F5E3")
            status_cell.font = Font(bold=True, color="1E8449")
        else:
            status_cell.fill = PatternFill("solid", fgColor="FADBD8")
            status_cell.font = Font(bold=True, color="C0392B")
        for col in range(1, 6):
            ws_tests.cell(row=row, column=col).border = border()
            ws_tests.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    col_widths = [10, 15, 45, 10, 14]
    for i, w in enumerate(col_widths, 1):
        ws_tests.column_dimensions[get_column_letter(i)].width = w

    # ── Feuille 3 : Bug Report ──────────────────────────────────────────────────
    ws_bugs = wb.create_sheet("🐛 Bug Report")
    bug_headers = ["Bug ID", "Test ID", "Module", "Severity", "Title", "Steps / Details", "Recommendation", "Reported On"]
    ws_bugs.append(bug_headers)

    for col in range(1, 9):
        cell = ws_bugs.cell(row=1, column=col)
        cell.fill = red
        cell.font = bold_white
        cell.alignment = center
        cell.border = border()

    severity_colors = {
        "Critical": "C0392B",
        "High":     "E67E22",
        "Medium":   "F1C40F",
        "Low":      "2980B9",
    }

    bug_id = 1
    for t in SIMULATED_TEST_RESULTS:
        if t["status"] == "FAIL":
            severity = SEVERITIES.get(t["module"], "Medium")
            title, recommendation = BUG_DETAILS.get(t["id"], (t["description"], "To be investigated"))
            reported_on = (datetime.now() - timedelta(days=random.randint(0, 5))).strftime("%Y-%m-%d")
            ws_bugs.append([
                f"BUG-{bug_id:03d}",
                t["id"],
                t["module"],
                severity,
                title,
                f"Run test {t['id']} — {t['description']}",
                recommendation,
                reported_on
            ])
            row = ws_bugs.max_row
            sev_cell = ws_bugs.cell(row=row, column=4)
            sev_cell.font = Font(bold=True, color=severity_colors.get(severity, "000000"))
            for col in range(1, 9):
                ws_bugs.cell(row=row, column=col).border = border()
                ws_bugs.cell(row=row, column=col).alignment = Alignment(horizontal="center", wrap_text=True)
            logger.warning(f"BUG-{bug_id:03d} | [{severity}] {t['module']} — {title}")
            bug_id += 1

    bug_col_widths = [10, 10, 14, 12, 45, 40, 45, 14]
    for i, w in enumerate(bug_col_widths, 1):
        ws_bugs.column_dimensions[get_column_letter(i)].width = w
    for row in ws_bugs.iter_rows(min_row=2):
        ws_bugs.row_dimensions[row[0].row].height = 40

    # ── Sauvegarde ─────────────────────────────────────────────────────────────
    filename = f"reports/bug_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    logger.info(f"✅ Report saved : {filename}")
    logger.info("=" * 60)
    print(f"\n🎉 Done! Report generated: {filename}")
    return filename


if __name__ == "__main__":
    generate_bug_report()
